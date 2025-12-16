from flask import Flask, jsonify, request, render_template
from flask_cors import CORS
import sqlite3
import os
import pandas as pd
from flask import send_file
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import tempfile

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.path.join(BASE_DIR, "tarefas.db")


# ===============================
# BANCO
# ===============================

def conectar():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def criar_banco():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS topicos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tarefas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            topico_id INTEGER,
            texto TEXT,
            urgencia TEXT,
            concluida INTEGER DEFAULT 0
        )
    """)

    conn.commit()
    conn.close()


criar_banco()


# ===============================
# FRONTEND
# ===============================

@app.route("/")
def index():
    return render_template("index.html")


# ===============================
# API
# ===============================

@app.route("/dados", methods=["GET"])
def carregar_dados():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM topicos")
    topicos_db = cursor.fetchall()

    dados = []

    for topico in topicos_db:
        cursor.execute("""
            SELECT * FROM tarefas
            WHERE topico_id = ?
        """, (topico["id"],))

        tarefas_db = cursor.fetchall()

        tarefas = [{
            "id": t["id"],
            "texto": t["texto"],
            "urgencia": t["urgencia"],
            "concluida": bool(t["concluida"])
        } for t in tarefas_db]

        dados.append({
            "id": topico["id"],
            "nome": topico["nome"],
            "tarefas": tarefas
        })

    conn.close()
    return jsonify(dados)


@app.route("/topico", methods=["POST"])
def criar_topico():
    data = request.get_json()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(
        "INSERT INTO topicos (nome) VALUES (?)",
        (data["nome"],)
    )

    topico_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({
        "id": topico_id,
        "nome": data["nome"],
        "tarefas": []
    })


@app.route("/tarefa", methods=["POST"])
def criar_tarefa():
    data = request.get_json()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO tarefas (topico_id, texto, urgencia, concluida)
        VALUES (?, ?, ?, 0)
    """, (
        data["topico_id"],
        data["texto"],
        data["urgencia"]
    ))

    tarefa_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({
        "id": tarefa_id,
        "texto": data["texto"],
        "urgencia": data["urgencia"],
        "concluida": False
    })


@app.route("/concluir", methods=["POST"])
def concluir_tarefa():
    data = request.get_json()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE tarefas
        SET concluida = 1
        WHERE id = ?
    """, (data["tarefa_id"],))

    conn.commit()
    conn.close()

    return jsonify({"status": "ok"})

@app.route("/exportar", methods=["GET"])
def exportar_excel():
    conn = conectar()

    query = """
        SELECT
            t.nome AS Topico,
            f.texto AS Tarefa,
            f.urgencia AS Urgencia,
            CASE f.concluida
                WHEN 1 THEN 'Concluída'
                ELSE 'Pendente'
            END AS Status
        FROM tarefas f
        JOIN topicos t ON t.id = f.topico_id
        ORDER BY t.nome, f.urgencia
    """

    df = pd.read_sql(query, conn)
    conn.close()

    if df.empty:
        df = pd.DataFrame(
            columns=["Topico", "Tarefa", "Urgencia", "Status"]
        )

    # ===============================
    # ARQUIVO TEMPORÁRIO (OPÇÃO 1)
    # ===============================
    nome_arquivo = f"relatorio_tarefas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho = os.path.join(tempfile.gettempdir(), nome_arquivo)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tarefas")
        ws = writer.book["Tarefas"]

        # ===============================
        # TABELA
        # ===============================
        tabela = Table(
            displayName="TabelaTarefas",
            ref=f"A1:D{len(df) + 1}"
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

        # ===============================
        # CABEÇALHO (FONTE BRANCA)
        # ===============================
        fonte_branca = Font(color="FFFFFF", bold=True)
        fundo_cabecalho = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = fonte_branca
            cell.fill = fundo_cabecalho

        # ===============================
        # AJUSTE DE COLUNAS
        # ===============================
        for coluna in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in coluna
            )
            ws.column_dimensions[coluna[0].column_letter].width = max_length + 3

    # ===============================
    # DOWNLOAD
    # ===============================
    return send_file(
        caminho,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=True)
